import pandas as pd
import math
from datetime import datetime

# Get the current date and time
now = datetime.now()

# Format as a string
#timestamp = now.strftime("%Y%m%d_%H%M%S")
timestamp = now.strftime("%Y%m%d")

# Create a filename with the timestamp
xlsx_filename = f"AnalysePileLogicielleDOE_ExaMA_{timestamp}.xlsx"


# Specify the path to your Excel file
file_path = 'doe.xlsx'

# Read the Excel spreadsheet
df = pd.read_excel(file_path)

# Print the DataFrame
print(df)

solvers_list = [
    "KokkosKernels",
    "PETSc",
    "PARDISO",
    "Trilinos",
    "SuperLU-Dist",
    "STRUMPACK",
    "Hypre",
    "SPARSKIT",
    "BLAS",
    "SuperLU",
    "SparsePACK",
    "LAPACK",
    "Krino",
    "Scipy",
    "Eigen",
    "MFEM Solvers",
    "Sundials",
    "Zoltan/Zoltan2",
    "ARPACK",
    "PyMatLib"
]
math_list = ["SAMRAI", "STK", "MFEM", "UMR", "Portage", "Tangram", "Axom", "Overlink", "METIS", "ParMETIS", "Sculpt", "libigl"] 
compilers_list = ["Kokkos", "RAJA Suite", "FleCSI", "Flang", "MPICH", "OpenMPI", "Legion", "PyKokkos", "KokkosRemoteMemorySpaces", "Fortran", "MPI", "C", "C++", "GCC", "HIP", "CUDA", "Python", "OpenMP", "Intel Compiler Suite", "LLVM", "Perl", "PyTorch", "TensorFlow", "Boost", "Intel MPI"]
system_list = ["CharlieCloud", "LDMS", "Flux", "SICM", "AppSysFusion", "GMI", "Maestro/Merlin", "Splunk", "SLURM", "VmWare", "LSF"]
visu_list = ["VTK/VTKm", "Paraview", "Visit", "Catalyst", "Conduit", "Cinema", "Ascent"]
build_list = ["Spack", "BLT", "CMake", "Ninja", "Autoconf/Automake", "gdb", "git", "Gitlab", "git-lfs", "Valgrind", "AllineaForge", "TotalView", "Caliper", "Archer", "PAPI", "KokkosTools", "CDash", "STAT"]
io_list = ["HDF5/Parallel-HDF5", "NetCDF, pNetCDF", "SEACAS", "UnifyFS", "ZFP", "HPSS, MarFS, SILO", "Exodus, yamlcpp", "GUFI, HIO, SCR, Sina/Kosh", "CGNS, libz", "DB2, Matio", "ADIOS, szip/AEC"]


# Define the mapping from scale to numerical values
scale_mapping = {
    "Low": 0,
    "Medium": 1,
    "High": 2,
    "Very High": 3
}

data_map = {
    "solvers_list": (["KokkosKernels", "PETSc", "PARDISO", "Trilinos", "SuperLU-Dist", "STRUMPACK", "Hypre", "SPARSKIT", "BLAS", "SuperLU", "SparsePACK", "LAPACK","Krino",    "Scipy",    "Eigen",    "MFEM Solvers",    "Sundials",    "Zoltan/Zoltan2",    "ARPACK",    "PyMatLib"], "Solver Libraries"),
    "math_list": (["SAMRAI", "STK", "MFEM", "UMR", "Portage", "Tangram", "Axom", "Overlink", "METIS", "ParMETIS", "Sculpt", "libigl"], "Math, Meshing, Discretization"),
    "compilers_list": (["Kokkos", "RAJA Suite", "FleCSI", "Flang", "MPICH", "OpenMPI", "Legion", "PyKokkos", "KokkosRemoteMemorySpaces", "Fortran", "MPI", "C", "C++", "GCC", "HIP", "CUDA", "Python", "OpenMP", "Intel Compiler Suite", "LLVM", "Perl", "PyTorch", "TensorFlow", "Boost", "Intel MPI"], "Compilers, Runtimes, Languages"),
    "system_list": (["CharlieCloud", "LDMS", "Flux", "SICM", "AppSysFusion", "GMI", "Maestro/Merlin", "Splunk", "SLURM", "VmWare", "LSF"], "System Imaging, Monitoring"),
    "visu_list": (["VTK/VTKm", "Paraview", "Visit", "Catalyst", "Conduit", "Cinema", "Ascent"], "Visualisation And Analysis"),
    "build_list": (["Spack", "BLT", "CMake", "Ninja", "Autoconf/Automake", "gdb", "git", "Gitlab", "git-lfs", "Valgrind", "AllineaForge", "TotalView", "Caliper", "Archer", "PAPI", "KokkosTools", "CDash", "STAT"], "Build, Development, Software"),
    "io_list": (["HDF5/Parallel-HDF5", "NetCDF, pNetCDF", "SEACAS", "UnifyFS", "ZFP", "HPSS, MarFS, SILO", "Exodus, yamlcpp", "GUFI, HIO, SCR, Sina/Kosh", "CGNS, libz", "DB2, Matio", "ADIOS, szip/AEC"], "IO Storage, Data Management")
}

# create a Pandas Excel writer using XlsxWriter as the engine
writer = pd.ExcelWriter(xlsx_filename, engine='xlsxwriter')

for software_list, title in data_map.values():
    print(f"Processing {title}: {software_list}")
    average_value = dict()
    # Define the reverse mapping
    reverse_mapping = {v: k for k, v in scale_mapping.items()}

    for software in software_list:
        software_columns = [column for column in df.columns if software in column]
        if software_columns:
            # Apply the mapping to the specific columns
            df[software_columns] = df[software_columns].replace(scale_mapping)
            # Ignore rows with all NaN values
            average_value[software] = df[software_columns].dropna(how='all').mean(axis=0)
            # Convert the average value back to the scale
            average_value[software] = average_value[software].round().map(reverse_mapping)
            print(f"Average value for {software}: {average_value[software]}")
            #print(f"Average value for {software}: {average_value[software]}")
            # Print the results for each column
            #for column in software_columns:
            #    print(f"Results for {column}:")
            #    print(df[column].map(reverse_mapping).dropna())
        else:
            print(f"No columns found for {software}")


    # Define the table structure
    table = {
        ("Low", "Low"): [],
        ("Low", "Medium"): [],
        ("Low", "High"): [],
        ("Low", "Very High"): [],
        ("Medium", "Low"): [],
        ("Medium", "Medium"): [],
        ("Medium", "High"): [],
        ("Medium", "Very High"): [],
        ("High", "Low"): [],
        ("High", "Medium"): [],
        ("High", "High"): [],
        ("High", "Very High"): [],
        ("Very High", "Low"): [],
        ("Very High", "Medium"): [],
        ("Very High", "High"): [],
        ("Very High", "Very High"): [],
    }

    # Fill the table with software names
    for software in software_list:
        row=dict()
        col=dict()
        ok=0
        for key, val in average_value[software].items():
            if "Likelihood" in key:
                if val in ["Very High","High","Medium","Low"]:
                    row[software]=val
                else:
                    row[software]="Low"
            if "Impact" in key:
                if val in ["Very High","High","Medium","Low"]:
                    col[software]=val
                else:
                    col[software]="Low"
        table[(row[software], col[software])].append(software)
#   print(table)        

    ## Create the AsciiDoc table
    asciidoc_table = f"""
.Likelihood-Impact Matrix for {title}
|===\n|Likelihood \\ Impact|Very High|High|Medium|Low\n
"""
    #
    for likelihood in ["Very High","High","Medium","Low"]:
        asciidoc_table += f"|{likelihood}"
        for impact in ["Very High","High","Medium","Low"]:
            asciidoc_table += f"|{', '.join(table[(likelihood, impact)])}"
        asciidoc_table += "\n"

    asciidoc_table += "|===\n\n"

    with open(f"doe-analysis.adoc", "a") as f:
        f.write(f"== {title}")
        f.write(asciidoc_table)

    # Impact and Likelihood levels
    levels = ["Very High","High","Medium","Low"]


    # Convert lists in table dictionary to comma separated strings
    for key, value in table.items():
        if value:  # if value is a non-empty list
            table[key] = ', '.join(value)
        else:  # if value is an empty list or NaN
            table[key] = ''
    print(table)
    # Create a DataFrame from the table dictionary
    table_dict = {(likelihood, impact): table.get((likelihood, impact), '')
                  for likelihood in levels for impact in levels}
    table_df = pd.DataFrame(table_dict, index=[0])

    # Transpose the DataFrame so that likelihoods are rows and impacts are columns
    table_df = table_df.transpose()
    table_df.index.names = ['Likelihood', 'Impact']
    table_df.columns = ['Software']
    table_df = table_df.unstack()

    # Replace MultiIndex by getting the 'Impact' level
    table_df.columns = table_df.columns.get_level_values('Impact')

    # Specify categorical data type for the DataFrame index and columns to maintain order
    table_df.columns = pd.CategoricalIndex(table_df.columns, categories=levels[::-1], ordered=True)
    table_df.index = pd.CategoricalIndex(table_df.index, categories=levels[::-1], ordered=True)
    print(f"1: index {table_df.index}")
    print(f"2: cols {table_df.columns}")

    # Sort the DataFrame index and columns
    table_df.sort_index(axis=0, inplace=True, ascending=False)
    table_df.sort_index(axis=1, inplace=True, ascending=False)


    print(table_df)
    # Clean the title to be a valid Excel sheet name
    cleaned_title = title.replace('/', ' ')[:31]

    # Write the DataFrame to the Excel file with a sheet for each software category
    table_df.to_excel(writer, sheet_name=cleaned_title)
    

writer._save()
writer.close()

import openpyxl

from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Color

# Load the workbook
book = openpyxl.load_workbook(xlsx_filename)

# Define the fill style for "Very High" cells
red_fill = PatternFill(fill_type='solid', fgColor='FFFF8080')  # Light Red 3
orange_fill = PatternFill(fill_type='solid', fgColor='FFFFC78C')  # light Orange 3
green_fill = PatternFill(fill_type='solid', fgColor='FFC6EFCE')  # light Green 3
black_fill = PatternFill(fill_type='solid', fgColor='FF000000')  # black
grey_fill = PatternFill(fill_type='solid', fgColor='FF808080')  # grey
anthracite_fill = PatternFill(fill_type='solid', fgColor='FF333333')  # anthracite
very_high_fill= PatternFill(fill_type='solid', fgColor='FF2F75B6')  # 
high_fill= PatternFill(fill_type='solid', fgColor='FF5b9bd5')  # 
medium_fill = PatternFill(fill_type='solid', fgColor='FF9dc3e6')  #
low_fill = PatternFill(fill_type='solid', fgColor='FFbdd7ee')  #

# Loop over each sheet in the workbook
for sheetname in book.sheetnames:

    # Access the sheet
    sheet = book[sheetname]


    # Clear the content in cell A1
    sheet['A1'].value = None

    # Shift existing rows down
    sheet.insert_rows(1)
    # Shift existing columns to the right
    sheet.insert_cols(1)

    sheet.merge_cells('A1:B2')
    sheet['A1'].value = sheetname
    sheet['A1'].fill = grey_fill
    sheet['A1'].font = Font(color="FFFFFFFF", bold=True,size=22)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Merge cells and add the text "Impact"
    sheet.merge_cells('C1:F1')
    sheet['C1'].value = 'Impact of not having software product/tool/library available'
    sheet['C1'].fill = black_fill
    sheet['C1'].font = Font(color="FFFFFFFF", bold=True,size=20)

    # Center align the text in the merged cell
    sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

    # Merge cells and add the text "Likelihood of Risk"
    sheet.merge_cells('A3:A6')
    sheet['A3'].value = 'Likelihood of Risk'

    # Center align the text in the merged cell
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center',text_rotation=90)
    sheet['A3'].fill = black_fill
    sheet['A3'].font = Font(color="FFFFFFFF", bold=True,size=20)

    sheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B3'].fill = very_high_fill
    sheet['B3'].font = Font(color="FFFFFFFF", bold=True,size=16)
    sheet['C2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['C2'].fill = very_high_fill
    sheet['C2'].font = Font(color="FFFFFFFF", bold=True,size=16)
    sheet['B4'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B4'].fill = high_fill
    sheet['B4'].font = Font(color="FFFFFFFF", bold=True,size=16)
    sheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['D2'].fill = high_fill
    sheet['D2'].font = Font(color="FFFFFFFF", bold=True,size=16)
    sheet['B5'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B5'].fill = medium_fill
    sheet['B5'].font = Font(color="FF000000", bold=True,size=16)
    sheet['E2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['E2'].fill = medium_fill
    sheet['E2'].font = Font(color="FF000000", bold=True,size=16)
    sheet['B6'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B6'].fill = low_fill
    sheet['B6'].font = Font(color="FF000000", bold=True,size=16)
    sheet['F2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F2'].fill = low_fill
    sheet['F2'].font = Font(color="FF000000", bold=True,size=16)

    # Set column widths
    for i, column in enumerate(sheet.columns):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 30 # set width to 120

   # Set row heights
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 100  # Set a specific height

    # Check conditions and apply fill to specified cells
    for row in sheet.iter_rows(min_row=3, max_row=6, min_col=3, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(color="FF000000",size=16)
            if cell.column  + cell.row >= 6 and cell.column  + cell.row <= 7:
                cell.fill = red_fill
            if (cell.column  + cell.row >= 8) and (cell.column  + cell.row <= 9):
                cell.fill = orange_fill
            if (cell.column  + cell.row >= 10) and (cell.column  + cell.row <= 12):
                cell.fill = green_fill

    side = Side(border_style='thick', color=Color('FFFFFFFF'))
    for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(left=side, right=side, top=side, bottom=side)


# Save the changes
book.save(xlsx_filename)